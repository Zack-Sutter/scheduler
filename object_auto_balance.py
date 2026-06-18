import copy
import datetime
import os
import random
import re
import subprocess
import sys
from dataclasses import dataclass

import numpy as np
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PySide6.QtCore import (
    QAbstractTableModel,
    QEasingCurve,
    QModelIndex,
    QPoint,
    QPropertyAnimation,
    QRect,
    Qt,
    QTimer,
    Signal,
)
from PySide6.QtGui import (
    QBrush,
    QColor,
    QCloseEvent,
    QEnterEvent,
    QFont,
    QFontMetrics,
    QKeySequence,
    QPainter,
    QPen,
)
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QDialog,
    QDialogButtonBox,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QSizePolicy,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTableView,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from object_auto_balance_core import (
    NONSTANDARD_SHIFTS,
    RES_FILE_NAME,
    SHIFT_INFO,
    STANDARD_FLOOR_SHIFTS,
    SWAPPABLE_FLOOR_SHIFTS,
    count_column_violations,
    default_balance_rules,
    direct_swap_blocked,
    format_balance_rule_line,
    is_standard_shift_covered,
    shift_background_color,
    total_violations,
)

SHIFT_COVERAGE_ROLE = Qt.ItemDataRole.UserRole + 1
SHIFT_COVERAGE_INDICATOR_WIDTH = 6

primary_button_color = '#EDD863'
primary_button_hover_color = '#E1D591'
secondary_button_color = '#d15462'
secondary_button_hover_color = '#ea6473'
enable_button_color = '#2e7d32'
enable_button_hover_color = '#388e3c'
disable_button_color = secondary_button_color
disable_button_hover_color = secondary_button_hover_color
background_color = '#ffffff'
frame_color = '#eeeeee'
text_color = '#000000'
subtitle_text_color = '#666666'
text_field_color = '#c6c6c6'
sheet_grid_line_color = '#b0b0b0'
regular_font_family = "Georgia"
regular_font_size = 11

APP_STYLESHEET = f"""
* {{
    color: {text_color};
    font-family: {regular_font_family};
    font-size: {regular_font_size}pt;
}}
QMainWindow, QWidget#centralRoot, QDialog {{
    background-color: {background_color};
}}
QFrame {{
    background-color: {frame_color};
}}
QLineEdit, QTextEdit {{
    background-color: {text_field_color};
    color: {text_color};
    font-size: {regular_font_size-1}pt;
}}
QListWidget, QTableView {{
    background-color: {text_field_color};
}}
QHeaderView::section {{
    background-color: {frame_color};
    color: {text_color};
}}
QPushButton#primaryBtn {{
    background-color: {primary_button_color};
    color: {text_color};
}}
QPushButton#primaryBtn:hover {{
    background-color: {primary_button_hover_color};
}}
QPushButton#secondaryBtn {{
    background-color: {secondary_button_color};
    color: {text_color};
}}
QPushButton#secondaryBtn:hover {{
    background-color: {secondary_button_hover_color};
}}
QPushButton#enableBtn {{
    background-color: {enable_button_color};
    color: {text_color};
}}
QPushButton#enableBtn:hover {{
    background-color: {enable_button_hover_color};
}}
QPushButton#disableBtn {{
    background-color: {disable_button_color};
    color: {text_color};
}}
QPushButton#disableBtn:hover {{
    background-color: {disable_button_hover_color};
}}
"""


class ToastManager:
    MARGIN = 16
    SLIDE_MS = 250
    DEFAULT_DURATION_MS = 2000

    def __init__(self, parent: QWidget):
        self._parent = parent
        self._active_toast: QWidget | None = None
        self._active_animation: QPropertyAnimation | None = None

    def show(self, message: str, background_color: str = enable_button_color, duration_ms: int = DEFAULT_DURATION_MS) -> None:
        if self._active_toast is not None:
            self._active_toast.deleteLater()
            self._active_toast = None
            self._active_animation = None

        toast = QWidget(self._parent)
        toast.setObjectName('toast')
        toast.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        toast.setStyleSheet(f"""
            QWidget#toast {{
                background-color: {background_color};
                border-radius: 6px;
            }}
        """)
        layout = QHBoxLayout(toast)
        layout.setContentsMargins(12, 8, 12, 8)
        label = QLabel(message)
        label.setStyleSheet(f'background-color: {background_color};')
        layout.addWidget(label)
        toast.adjustSize()

        parent_width = self._parent.width()
        x = parent_width - toast.width() - self.MARGIN
        y = self.MARGIN
        toast.move(parent_width, y)
        toast.show()
        toast.raise_()

        self._active_toast = toast
        self._slide_to(
            toast,
            x,
            y,
            on_finished=lambda: QTimer.singleShot(
                duration_ms, lambda: self._dismiss(toast)
            ),
        )

    def _slide_to(
        self,
        toast: QWidget,
        x: int,
        y: int,
        on_finished=None,
    ) -> None:
        animation = QPropertyAnimation(toast, b'pos')
        animation.setDuration(self.SLIDE_MS)
        animation.setStartValue(toast.pos())
        animation.setEndValue(QPoint(x, y))
        animation.setEasingCurve(QEasingCurve.Type.OutCubic)
        if on_finished is not None:
            animation.finished.connect(on_finished)
        animation.start()
        self._active_animation = animation

    def _dismiss(self, toast: QWidget) -> None:
        if self._active_toast is not toast:
            return
        animation = QPropertyAnimation(toast, b'pos')
        animation.setDuration(self.SLIDE_MS)
        animation.setStartValue(toast.pos())
        animation.setEndValue(QPoint(self._parent.width(), toast.y()))
        animation.setEasingCurve(QEasingCurve.Type.InCubic)
        animation.finished.connect(toast.deleteLater)
        animation.finished.connect(self._clear_active_toast)
        animation.start()
        self._active_animation = animation

    def _clear_active_toast(self) -> None:
        self._active_toast = None
        self._active_animation = None


info_icon_color = subtitle_text_color
info_tip_background_color = '#ffffff'
info_tip_border_color = '#b0b0b0'


class InfoTipManager:
    SHOW_DELAY_MS = 400
    MAX_TIP_WIDTH = 300

    def __init__(self, parent: QWidget):
        self._parent = parent
        self._popup: QWidget | None = None
        self._label: QLabel | None = None
        self._anchor: QWidget | None = None
        self._pending_text = ''
        self._show_timer = QTimer()
        self._show_timer.setSingleShot(True)
        self._show_timer.timeout.connect(self._display_popup)

    def create_icon(self, text: str, parent: QWidget | None = None) -> 'InfoIcon':
        return InfoIcon(text, self, parent)

    def request_show(self, icon: 'InfoIcon', text: str) -> None:
        self._anchor = icon
        self._pending_text = text
        self._show_timer.start(self.SHOW_DELAY_MS)

    def request_hide(self, icon: 'InfoIcon') -> None:
        if self._anchor is icon:
            self._show_timer.stop()
            self._hide_popup()

    def _ensure_popup(self) -> None:
        if self._popup is not None:
            return
        popup = QFrame(self._parent, Qt.WindowType.ToolTip)
        popup.setObjectName('infoTipPopup')
        popup.setStyleSheet(f"""
            QFrame#infoTipPopup {{
                background-color: {info_tip_background_color};
                border: 1px solid {info_tip_border_color};
                border-radius: 4px;
            }}
        """)
        layout = QHBoxLayout(popup)
        layout.setContentsMargins(8, 6, 8, 6)
        label = QLabel()
        label.setWordWrap(True)
        label.setMaximumWidth(self.MAX_TIP_WIDTH)
        layout.addWidget(label)
        self._popup = popup
        self._label = label

    def _display_popup(self) -> None:
        if self._anchor is None or not self._pending_text:
            return
        self._ensure_popup()
        assert self._popup is not None and self._label is not None
        self._label.setText(self._pending_text)
        self._popup.adjustSize()

        anchor = self._anchor
        global_pos = anchor.mapToGlobal(QPoint(0, anchor.height()))
        popup = self._popup
        popup.move(global_pos)
        popup.show()
        popup.raise_()

    def _hide_popup(self) -> None:
        if self._popup is not None:
            self._popup.hide()
        self._anchor = None
        self._pending_text = ''


class InfoIcon(QLabel):
    ICON_SIZE = 16

    def __init__(self, text: str, manager: InfoTipManager, parent: QWidget | None = None):
        super().__init__('i', parent)
        self._tip_text = text
        self._manager = manager
        self.setObjectName('infoIcon')
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setCursor(Qt.CursorShape.WhatsThisCursor)
        self.setFixedSize(self.ICON_SIZE, self.ICON_SIZE)
        self.setStyleSheet(f"""
            QLabel#infoIcon {{
                color: {info_icon_color};
                background-color: transparent;
                border: 1px solid {info_icon_color};
                border-radius: {self.ICON_SIZE // 2}px;
                font-size: {max(regular_font_size - 2, 8)}pt;
                font-weight: bold;
            }}
            QLabel#infoIcon:hover {{
                color: {text_color};
                border-color: {text_color};
            }}
        """)

    def enterEvent(self, event: QEnterEvent) -> None:
        self._manager.request_show(self, self._tip_text)
        super().enterEvent(event)

    def leaveEvent(self, event) -> None:
        self._manager.request_hide(self)
        super().leaveEvent(event)


def label_with_subtitle(title: str, subtitle: str) -> QWidget:
    container = QWidget()
    layout = QVBoxLayout(container)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.setSpacing(0)
    title_label = QLabel(title)
    title_font = title_label.font()
    title_font.setPointSize(regular_font_size)
    title_label.setFont(title_font)
    layout.addWidget(title_label)
    sub = QLabel(subtitle)
    sub.setWordWrap(True)
    sub_font = sub.font()
    sub_font.setPointSize(max(regular_font_size - 2, 6))
    sub.setFont(sub_font)
    sub.setStyleSheet(f'color: {subtitle_text_color};')
    layout.addWidget(sub)
    return container


@dataclass
class SelectionRegion:
    from_row: int
    upto_row: int
    from_col: int
    upto_col: int

    @property
    def row_count(self) -> int:
        return self.upto_row - self.from_row

    @property
    def col_count(self) -> int:
        return self.upto_col - self.from_col

    def time_range(self, df: pd.DataFrame) -> tuple:
        return df.index[self.from_row], df.index[self.upto_row - 1]

    def column_names(self, df: pd.DataFrame) -> list:
        return list(df.columns[self.from_col : self.upto_col])

    def values_2d(self, df: pd.DataFrame) -> list:
        block = df.iloc[self.from_row : self.upto_row, self.from_col : self.upto_col]
        return block.values.tolist()

    def values_flat(self, df: pd.DataFrame) -> list:
        flat = []
        for row in self.values_2d(df):
            for cell in row:
                flat.append(cell)
        return flat


def _cell_to_clipboard_text(value) -> str:
    if pd.isna(value) or value == '':
        return ''
    return str(value)


def selection_values_to_tsv(values_2d: list[list]) -> str:
    return '\n'.join(
        '\t'.join(_cell_to_clipboard_text(cell) for cell in row) for row in values_2d
    )


def parse_clipboard_tsv(text: str) -> list[list[str]]:
    if not text:
        return []
    lines = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    if lines and lines[-1] == '':
        lines = lines[:-1]
    return [line.split('\t') for line in lines]


def apply_sheet_paste(
    df: pd.DataFrame, region: SelectionRegion, grid: list[list[str]]
) -> None:
    if not grid:
        return
    grid_rows = len(grid)
    grid_cols = max(len(row) for row in grid)
    if grid_cols == 0:
        return

    paste_rows = min(grid_rows, len(df.index) - region.from_row)
    paste_cols = min(grid_cols, len(df.columns) - region.from_col)
    if paste_rows <= 0 or paste_cols <= 0:
        return

    row_labels = df.index[region.from_row : region.from_row + paste_rows]
    col_labels = df.columns[region.from_col : region.from_col + paste_cols]
    values = []
    for row_idx in range(paste_rows):
        row = grid[row_idx]
        row_vals = []
        for col_idx in range(paste_cols):
            text = row[col_idx] if col_idx < len(row) else ''
            row_vals.append(np.nan if text == '' else text)
        values.append(row_vals)

    block = pd.DataFrame(values, index=row_labels, columns=col_labels)
    df.loc[row_labels, col_labels] = block


class ScheduleTableModel(QAbstractTableModel):
    EMPTY_COUNT_COL = 0

    def __init__(self, parent=None):
        super().__init__(parent)
        self._df = pd.DataFrame()

    @classmethod
    def df_column_index(cls, model_col: int) -> int | None:
        if model_col == cls.EMPTY_COUNT_COL:
            return None
        return model_col - 1

    @classmethod
    def model_column_index(cls, df_col: int) -> int:
        return df_col + 1

    def _row_empty_count(self, row: int) -> int:
        return sum(1 for value in self._df.iloc[row] if pd.isna(value) or value == '')

    def set_dataframe(self, df: pd.DataFrame, full_reset: bool = False) -> None:
        if full_reset or self._df.shape != df.shape or list(self._df.columns) != list(df.columns):
            self.beginResetModel()
            self._df = df.copy()
            self.endResetModel()
        else:
            self._df = df.copy()
            top_left = self.index(0, 0)
            bottom_right = self.index(
                max(0, len(self._df) - 1),
                max(0, len(self._df.columns)),
            )
            self.dataChanged.emit(
                top_left, bottom_right, [Qt.DisplayRole, Qt.BackgroundRole, Qt.ForegroundRole]
            )

    def rowCount(self, parent=QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self._df)

    def columnCount(self, parent=QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self._df.columns) + 1

    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if index.column() == self.EMPTY_COUNT_COL:
            if role == Qt.DisplayRole:
                return str(self._row_empty_count(index.row()))
            if role == Qt.TextAlignmentRole:
                return Qt.AlignCenter
            if role == Qt.ForegroundRole:
                return QBrush(QColor(subtitle_text_color))
            return None
        value = self._df.iat[index.row(), index.column() - 1]
        if role == Qt.DisplayRole:
            return '' if pd.isna(value) else str(value)
        if role == Qt.BackgroundRole:
            color = shift_background_color(value)
            if color:
                return QBrush(QColor(color))
        if role == Qt.ForegroundRole:
            return QBrush(QColor(text_color))
        return None

    def headerData(self, section: int, orientation: Qt.Orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            if section == self.EMPTY_COUNT_COL:
                return ''
            df_col = section - 1
            if 0 <= df_col < len(self._df.columns):
                return str(self._df.columns[df_col])
        elif section < len(self._df.index):
            return str(self._df.index[section])
        return None

    def flags(self, index: QModelIndex):
        if not index.isValid():
            return Qt.NoItemFlags
        if index.column() == self.EMPTY_COUNT_COL:
            return Qt.ItemIsEnabled
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class ScheduleColumnHeader(QHeaderView):
    def __init__(self, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self._grid_color = QColor(sheet_grid_line_color)

    def paintSection(self, painter: QPainter, rect, logicalIndex: int) -> None:
        super().paintSection(painter, rect, logicalIndex)
        painter.save()
        painter.setPen(QPen(self._grid_color, 1))
        right = rect.right()
        bottom = rect.bottom()
        painter.drawLine(right, rect.top(), right, bottom)
        painter.drawLine(rect.left(), bottom, right, bottom)
        painter.restore()


class ScheduleTableView(QTableView):
    regions_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._regions: list[SelectionRegion] = []
        self._drag_anchor: tuple[int, int] | None = None
        self._drag_preview: SelectionRegion | None = None
        self.setSelectionMode(QTableView.NoSelection)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        column_header = ScheduleColumnHeader(self)
        self.setHorizontalHeader(column_header)
        column_header.sectionClicked.connect(self._on_column_header_clicked)

    def scrollContentsBy(self, dx: int, dy: int) -> None:
        pass

    def wheelEvent(self, event):
        event.ignore()

    def selection_regions(self) -> list[SelectionRegion]:
        return list(self._regions)

    def set_regions(self, regions: list[SelectionRegion]) -> None:
        self._regions = regions
        self.viewport().update()
        self.regions_changed.emit()

    def _df_col_range_from_model_cols(self, col_a: int, col_b: int) -> tuple[int, int] | None:
        lo, hi = min(col_a, col_b), max(col_a, col_b)
        df_cols = [
            c
            for c in (
                ScheduleTableModel.df_column_index(col)
                for col in range(lo, hi + 1)
            )
            if c is not None
        ]
        if not df_cols:
            return None
        return min(df_cols), max(df_cols) + 1

    def _region_from_model_cells(
        self, r0: int, c0: int, r1: int, c1: int
    ) -> SelectionRegion | None:
        col_range = self._df_col_range_from_model_cols(c0, c1)
        if col_range is None:
            return None
        from_col, upto_col = col_range
        return SelectionRegion(
            min(r0, r1),
            max(r0, r1) + 1,
            from_col,
            upto_col,
        )

    def _set_drag_preview(self, region: SelectionRegion | None) -> None:
        if region == self._drag_preview:
            return
        self._drag_preview = region
        self.viewport().update()

    def _on_column_header_clicked(self, column: int) -> None:
        if self.model() is None or self.model().rowCount() == 0:
            return
        if column == ScheduleTableModel.EMPTY_COUNT_COL:
            return
        df_col = ScheduleTableModel.df_column_index(column)
        if df_col is None:
            return
        region = SelectionRegion(0, self.model().rowCount(), df_col, df_col + 1)
        self.set_regions([region])

    def mousePressEvent(self, event):
        self._drag_preview = None
        index = self.indexAt(event.position().toPoint())
        if index.isValid():
            self._drag_anchor = (index.row(), index.column())
        else:
            self._drag_anchor = None
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_anchor is not None and event.buttons() & Qt.LeftButton:
            index = self.indexAt(event.position().toPoint())
            if index.isValid():
                r0, c0 = self._drag_anchor
                self._set_drag_preview(
                    self._region_from_model_cells(
                        r0, c0, index.row(), index.column()
                    )
                )
            else:
                self._set_drag_preview(None)
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._drag_anchor is not None:
            end_index = self.indexAt(event.position().toPoint())
            if end_index.isValid():
                r0, c0 = self._drag_anchor
                region = self._region_from_model_cells(
                    r0, c0, end_index.row(), end_index.column()
                )
                if region is not None:
                    if event.modifiers() & Qt.ControlModifier:
                        regions = self._regions + [region]
                        if len(regions) > 2:
                            regions = regions[-2:]
                        self.set_regions(regions)
                    else:
                        self.set_regions([region])
        self._drag_anchor = None
        self._drag_preview = None
        self.viewport().update()
        super().mouseReleaseEvent(event)

    def _schedule_controller(self):
        parent = self.parent()
        if parent is not None and hasattr(parent, 'controller'):
            return parent.controller
        return None

    def keyPressEvent(self, event):
        controller = self._schedule_controller()
        if controller is None:
            super().keyPressEvent(event)
            return
        if event.matches(QKeySequence.Copy):
            controller.copy_sheet_selection()
            event.accept()
            controller.statusBar().showMessage("Copied to clipboard", 2000)
            return
        if event.matches(QKeySequence.Paste):
            controller.paste_sheet_selection()
            event.accept()
            return
        if event.matches(QKeySequence.Cut):
            controller.cut_sheet_selection()
            event.accept()
            return
        if event.matches(QKeySequence.Undo):
            controller.undo()
            event.accept()
            return
        if event.matches(QKeySequence.Redo):
            controller.redo()
            event.accept()
            return
        if event.matches(QKeySequence.Bold):
            controller._perform_with_undo(controller.swap)
            event.accept()
            return
        super().keyPressEvent(event)

    def _paint_region_outline(self, painter: QPainter, region: SelectionRegion) -> None:
        top_left = self.visualRect(
            self.model().index(
                region.from_row, ScheduleTableModel.model_column_index(region.from_col)
            )
        )
        bottom_right = self.visualRect(
            self.model().index(
                region.upto_row - 1,
                ScheduleTableModel.model_column_index(region.upto_col - 1),
            )
        )
        rect = top_left.united(bottom_right)
        painter.drawRect(rect.adjusted(0, 0, -1, -1))

    def paintEvent(self, event):
        super().paintEvent(event)
        regions = list(self._regions)
        if self._drag_preview is not None:
            regions.append(self._drag_preview)
        if not regions:
            return

        painter = QPainter(self.viewport())
        painter.setPen(QPen(QColor('#2563eb'), 2))
        for region in regions:
            self._paint_region_outline(painter, region)
        painter.end()


class BalanceRulesDialog(QDialog):
    def __init__(self, parent, on_apply):
        super().__init__(parent)
        self.on_apply = on_apply
        self.setWindowTitle('Balance Shifts')
        self.setModal(True)
        self.setMinimumWidth(400)
        self.rules = copy.deepcopy(default_balance_rules())

        layout = QVBoxLayout(self)
        layout.addWidget(label_with_subtitle(
            'Rules run top to bottom',
            'Drag to reorder.  Double click to enable/disable',
        ))

        self.rule_list = QListWidget()
        self.rule_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.rule_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.rule_list.itemDoubleClicked.connect(self._on_rule_double_clicked)
        self.rule_list.model().rowsMoved.connect(self._on_rules_reordered)
        layout.addWidget(self.rule_list, stretch=1)

        buttons = QDialogButtonBox()
        cancel_btn = QPushButton('Cancel')
        cancel_btn.setObjectName('secondaryBtn')
        cancel_btn.clicked.connect(self.reject)
        apply_btn = QPushButton('Apply')
        apply_btn.setObjectName('primaryBtn')
        apply_btn.clicked.connect(self._apply)
        buttons.addButton(cancel_btn, QDialogButtonBox.RejectRole)
        buttons.addButton(apply_btn, QDialogButtonBox.AcceptRole)
        layout.addWidget(buttons)

        self.setStyleSheet(APP_STYLESHEET)
        self.refresh_rule_list()
        self.rule_list.setCurrentRow(0)

    def _style_rule_item(self, item: QListWidgetItem, rule) -> None:
        color = text_color if rule.enabled else subtitle_text_color
        item.setForeground(QBrush(QColor(color)))

    def refresh_rule_list(self):
        row = self.rule_list.currentRow()
        self.rule_list.clear()
        for index, rule in enumerate(self.rules):
            item = QListWidgetItem(format_balance_rule_line(index, rule))
            item.setData(Qt.ItemDataRole.UserRole, rule)
            self._style_rule_item(item, rule)
            self.rule_list.addItem(item)
        if self.rules:
            row = min(max(row, 0), len(self.rules) - 1)
            self.rule_list.setCurrentRow(row)

    def _sync_rules_from_list(self) -> None:
        self.rules = [
            self.rule_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self.rule_list.count())
        ]
        for index, rule in enumerate(self.rules):
            item = self.rule_list.item(index)
            item.setText(format_balance_rule_line(index, rule))
            self._style_rule_item(item, rule)

    def _on_rules_reordered(self, parent, start, end, destination, row) -> None:
        self._sync_rules_from_list()

    def _on_rule_double_clicked(self, item: QListWidgetItem) -> None:
        row = self.rule_list.row(item)
        if row < 0:
            return
        rule = item.data(Qt.ItemDataRole.UserRole)
        rule.enabled = not rule.enabled
        self.refresh_rule_list()
        self.rule_list.setCurrentRow(row)

    def _apply(self):
        self.on_apply(self.rules)
        self.accept()


class SheetFrame(QWidget):
    FRAME_WIDTH = 975
    FRAME_HEIGHT = 365

    def __init__(self, controller: 'ScheduleApp'):
        super().__init__()
        self.controller = controller
        self._frame_width = self.FRAME_WIDTH
        self._frame_height = self.FRAME_HEIGHT
        self._has_schedule = False
        self.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed
        )
        self.setMinimumWidth(self.FRAME_WIDTH)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.blank_panel = QWidget()
        self.blank_panel.setStyleSheet(f'background-color: {text_field_color};')
        layout.addWidget(self.blank_panel, stretch=1)
        self.model = ScheduleTableModel(self)
        self.table_view = ScheduleTableView(self)
        self.table_view.setObjectName('scheduleTable')
        self.table_view.setModel(self.model)
        self.table_view.setShowGrid(True)
        self.table_view.setCornerButtonEnabled(False)
        self.table_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table_view.setStyleSheet(
            f'QTableView#scheduleTable {{ gridline-color: {sheet_grid_line_color}; }}'
        )
        h_header = self.table_view.horizontalHeader()
        v_header = self.table_view.verticalHeader()
        h_header.setSectionResizeMode(QHeaderView.Fixed)
        h_header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        h_header.setMinimumSectionSize(1)
        h_header.setStretchLastSection(False)
        v_header.setSectionResizeMode(QHeaderView.Stretch)
        v_header.setMinimumSectionSize(1)
        v_header.setStretchLastSection(False)
        self.table_view.hide()
        layout.addWidget(self.table_view, stretch=1)
        self.schedule_info_icon = controller.info_tip_manager.create_icon(
            'Hotkeys\n'
            'Undo---Cmnd+Z\n'
            'Redo---Cmnd+Shift+Z\n'
            'Copy---Cmnd+C\n'
            'Paste---Cmnd+V\n'
            'Cut---Cmnd+X\n'
            'Swap---Cmnd+B\n',
            parent=self.table_view,
        )
        self.schedule_info_icon.hide()

    def has_schedule(self) -> bool:
        return self._has_schedule

    def show_schedule(self) -> None:
        self._has_schedule = True
        self.blank_panel.hide()
        self.table_view.show()
        self.schedule_info_icon.show()
        self._position_schedule_info_icon()

    def set_frame_size(self, width: int, height: int) -> None:
        self._frame_width = width
        self._frame_height = height
        self.setMinimumWidth(width)
        self.setFixedHeight(height)

    def showEvent(self, event):
        super().showEvent(event)
        self._fit_table_to_frame()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._fit_table_to_frame()

    def _frame_size(self) -> tuple[int, int]:
        width, height = self.width(), self.height()
        if width <= 1:
            width = self._frame_width
        if height <= 1:
            height = self._frame_height
        return width, height

    def _viewport_size(self) -> tuple[int, int]:
        frame_w, frame_h = self._frame_size()
        view = self.table_view
        h_header = view.horizontalHeader()
        v_header = view.verticalHeader()
        header_h = max(h_header.height(), h_header.sizeHint().height())
        header_w = max(v_header.width(), v_header.sizeHint().width())
        return max(1, frame_w - header_w), max(1, frame_h - header_h)

    def _fit_table_to_frame(self) -> None:
        if not self._has_schedule:
            return
        view = self.table_view
        rows = self.model.rowCount()
        cols = self.model.columnCount()
        if rows == 0 or cols == 0:
            return

        viewport_w = view.viewport().width()
        if viewport_w <= 1:
            viewport_w, _ = self._viewport_size()
        if view.showGrid():
            viewport_w = max(1, viewport_w - max(0, cols - 1))

        fm = QFontMetrics(view.font())
        empty_count_w = fm.horizontalAdvance('00') + 8
        data_cols = cols - 1
        data_viewport_w = max(1, viewport_w - empty_count_w)
        view.setColumnWidth(ScheduleTableModel.EMPTY_COUNT_COL, empty_count_w)
        base_col_w = data_viewport_w // data_cols
        extra_col_w = data_viewport_w - base_col_w * data_cols
        for offset, col in enumerate(range(1, cols)):
            view.setColumnWidth(col, base_col_w + (1 if offset < extra_col_w else 0))

        view.verticalHeader().resizeSections()
        view.horizontalScrollBar().setValue(0)
        view.verticalScrollBar().setValue(0)
        self._position_schedule_info_icon()

    def _position_schedule_info_icon(self) -> None:
        if not self._has_schedule:
            return
        view = self.table_view
        v_header = view.verticalHeader()
        h_header = view.horizontalHeader()
        corner_w = max(v_header.width(), 1)
        corner_h = max(h_header.height(), 1)
        icon_size = min(corner_w, corner_h, InfoIcon.ICON_SIZE + 4) - 2
        icon_size = max(icon_size, 12)
        icon = self.schedule_info_icon
        icon.setFixedSize(icon_size, icon_size)
        icon.move(1, 1)
        icon.raise_()

    def update_sheet(self):
        if not self._has_schedule:
            return
        df = self.controller.df.copy()
        full_reset = self.model.rowCount() == 0 or self.model.columnCount() == 0
        self.model.set_dataframe(df.fillna(''), full_reset=full_reset)
        self.table_view.set_regions([])
        self._fit_table_to_frame()
        QTimer.singleShot(0, self._fit_table_to_frame)

    def column_names(self) -> list:
        return list(self.controller.df.columns)


class InputFrame(QWidget):
    @staticmethod
    def _configure_action_button(button: QPushButton) -> None:
        button.setFixedHeight(button.sizeHint().height() * 2)
        button.setSizePolicy(
            button.sizePolicy().horizontalPolicy(),
            QSizePolicy.Policy.Fixed,
        )

    def __init__(self, controller: 'ScheduleApp'):
        super().__init__()
        self.controller = controller
        grid = QGridLayout(self)
        self.standard_frame = StandardShiftFrame(self, controller)
        self.nonstandard_frame = NonStandardShiftFrame(self, controller)
        grid.addWidget(self.standard_frame, 0, 0, 5, 1)
        grid.addWidget(self.nonstandard_frame, 0, 1, 5, 1)

        self.undo_button = QPushButton('Undo')
        self.undo_button.setObjectName('secondaryBtn')
        self.undo_button.clicked.connect(controller.undo)
        grid.addWidget(self.undo_button, 0, 2)

        self.redo_button = QPushButton('Redo')
        self.redo_button.setObjectName('secondaryBtn')
        self.redo_button.clicked.connect(controller.redo)
        grid.addWidget(self.redo_button, 0, 3)

        self.swap_button = QPushButton('Swap')
        self.swap_button.setObjectName('primaryBtn')
        self.swap_button.clicked.connect(
            lambda: controller._perform_with_undo(controller.swap)
        )
        grid.addWidget(self.swap_button, 1, 2, 1, 1)

        self.delete_column_button = QPushButton('Delete Column')
        self.delete_column_button.setObjectName('secondaryBtn')
        self.delete_column_button.clicked.connect(controller.delete_column_clicked)
        grid.addWidget(self.delete_column_button, 2, 2, 1, 1)

        self.balance_button = QPushButton('Balance')
        self.balance_button.setObjectName('primaryBtn')
        self.balance_button.clicked.connect(controller.show_balance_rules_dialog)
        grid.addWidget(self.balance_button, 1, 3, 1, 1)

    def configure_action_buttons(self) -> None:
        for action_button in (
            self.undo_button,
            self.redo_button,
            self.swap_button,
            self.delete_column_button,
            self.balance_button,
        ):
            self._configure_action_button(action_button)


class ShiftCoverageDelegate(QStyledItemDelegate):
    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index: QModelIndex) -> None:
        content_option = QStyleOptionViewItem(option)
        content_option.rect = option.rect.adjusted(
            0, 0, -SHIFT_COVERAGE_INDICATOR_WIDTH, 0
        )
        super().paint(painter, content_option, index)

        covered = index.data(SHIFT_COVERAGE_ROLE)
        if covered is None:
            return
        color = QColor(enable_button_color if covered else secondary_button_color)
        indicator = QRect(
            option.rect.right() - SHIFT_COVERAGE_INDICATOR_WIDTH + 1,
            option.rect.top() + 1,
            SHIFT_COVERAGE_INDICATOR_WIDTH - 1,
            option.rect.height() - 2,
        )
        painter.fillRect(indicator, color)


class NonStandardShiftFrame(QFrame):
    def __init__(self, parent: InputFrame, controller: 'ScheduleApp'):
        super().__init__(parent)
        self.controller = controller
        self.setFrameStyle(QFrame.Box | QFrame.Plain)
        grid = QGridLayout(self)
        grid.addWidget(
            label_with_subtitle('Add nonstandard shift', 'fill selection'), 0, 0, 1, 1
        )
        self.list_widget = QListWidget()
        self.list_widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_widget.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        fm = QFontMetrics(self.list_widget.font())
        list_width = fm.horizontalAdvance('0' * 20) + 2 * self.list_widget.frameWidth() + 8
        self.list_widget.setFixedWidth(list_width)
        for shift in [*NONSTANDARD_SHIFTS, 'DELETE']:
            item = QListWidgetItem(shift)
            color = shift_background_color(shift) if shift != 'DELETE' else '#D3D3D3'
            if color:
                item.setBackground(QBrush(QColor(color)))
            item.setForeground(QBrush(QColor(text_color)))
            self.list_widget.addItem(item)
        list_height = sum(
            self.list_widget.sizeHintForRow(row) for row in range(self.list_widget.count())
        )
        list_widget_height = list_height + 2 * self.list_widget.frameWidth()
        self.list_widget.setFixedHeight(list_widget_height)
        self.list_widget.itemClicked.connect(self.on_list_item_clicked)
        grid.addWidget(self.list_widget, 1, 0)

        self.entry = QLineEdit()
        self.entry.setPlaceholderText('Custom shift')
        self.entry.setFixedWidth(list_width)
        self.entry.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.entry.returnPressed.connect(self.add_custom_shift_action)
        grid.addWidget(self.entry, 2, 0, 1, 1)

    def _apply_shift_to_selection(self, shift):
        selection = self.controller.get_sheet_selection()
        if not selection:
            return
        selection['shift'] = shift
        self.controller._perform_with_undo(
            self.controller.add_nonstandard_shift, selection
        )

    def on_list_item_clicked(self, item: QListWidgetItem):
        shift = item.text()
        if shift == 'DELETE':
            shift = np.nan
        self._apply_shift_to_selection(shift)

    def add_custom_shift_action(self):
        shift = self.entry.text().strip()
        if not shift:
            return
        self._apply_shift_to_selection(shift)
        self.entry.clear()


class StandardShiftFrame(QFrame):
    def __init__(self, parent: InputFrame, controller: 'ScheduleApp'):
        super().__init__(parent)
        self.controller = controller
        self.setFrameStyle(QFrame.Box | QFrame.Plain)
        grid = QGridLayout(self)
        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(4, 4, 4, 0)
        header_layout.setSpacing(4)
        header_layout.addWidget(
            controller.info_tip_manager.create_icon('green bar = covered all day', parent=header),
            alignment=Qt.AlignmentFlag.AlignTop,
        )
        header_layout.addWidget(
            label_with_subtitle('Add standard shift', 'one item per row')
        )
        grid.addWidget(header, 0, 0)
        self.list_widget = QListWidget()
        self.list_widget.setItemDelegate(ShiftCoverageDelegate(self.list_widget))
        self.list_widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_widget.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        fm = QFontMetrics(self.list_widget.font())
        list_width = (
            fm.horizontalAdvance('0' * 20)
            + SHIFT_COVERAGE_INDICATOR_WIDTH
            + 2 * self.list_widget.frameWidth()
            + 8
        )
        self.list_widget.setFixedWidth(list_width)
        for shift in STANDARD_FLOOR_SHIFTS:
            item = QListWidgetItem(shift)
            color = shift_background_color(shift)
            if color:
                item.setBackground(QBrush(QColor(color)))
            item.setForeground(QBrush(QColor(text_color)))
            self.list_widget.addItem(item)
        list_height = sum(
            self.list_widget.sizeHintForRow(row) for row in range(self.list_widget.count())
        )
        list_widget_height = list_height + 2 * self.list_widget.frameWidth()
        self.list_widget.setFixedHeight(list_widget_height)
        self.list_widget.itemClicked.connect(self.on_list_item_clicked)
        grid.addWidget(self.list_widget, 1, 0)
        self.update_coverage_indicators()

    def update_coverage_indicators(self) -> None:
        has_schedule = self.controller.sheet_frame.has_schedule()
        df = self.controller.df
        for row in range(self.list_widget.count()):
            item = self.list_widget.item(row)
            if has_schedule and not df.empty:
                covered = is_standard_shift_covered(df, item.text())
            else:
                covered = False
            item.setData(SHIFT_COVERAGE_ROLE, covered)
        self.list_widget.viewport().update()

    def on_list_item_clicked(self, item: QListWidgetItem):
        self.controller._perform_with_undo(
            self.controller.add_standard_shift, item.text()
        )


class ScheduleApp(QMainWindow):
    LEFT_PANEL_MAX_WIDTH = 280

    def __init__(self):
        super().__init__()
        self.setWindowTitle('MoMath Automatic Scheduler')
        self.resize(1500, 800)
        self.setStyleSheet(APP_STYLESHEET)

        self.nonstandard_shifts = []
        self.action_history_stack = []
        self.action_redo_stack = []
        self.df = pd.DataFrame()
        self.info_tip_manager = InfoTipManager(self)
        self.sheet_frame = SheetFrame(self)
        self.sheet_frame.set_frame_size(
            SheetFrame.FRAME_WIDTH, SheetFrame.FRAME_HEIGHT
        )
        self.inputs: InputFrame | None = None
        self.paid_workers: list = []
        self.volunteers: list = []

        root = QWidget()
        root.setObjectName('centralRoot')
        self.setCentralWidget(root)
        main_layout = QHBoxLayout(root)

        left_panel = QWidget()
        left_panel.setMaximumWidth(self.LEFT_PANEL_MAX_WIDTH)
        left_panel.setSizePolicy(
            QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Expanding
        )
        left = QVBoxLayout(left_panel)
        left.setContentsMargins(6, 6, 6, 6)
        left.setSpacing(6)
        main_layout.addWidget(left_panel, stretch=0)

        paid_worker_header = label_with_subtitle('Paid workers', 'comma separated')
        paid_worker_header.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        left.addWidget(paid_worker_header)

        self.paid_workers_entry = QTextEdit()
        self.paid_workers_entry.setFixedHeight(60)
        left.addWidget(self.paid_workers_entry)

        volunteer_header = label_with_subtitle('Volunteers', 'comma separated')
        volunteer_header.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        left.addWidget(volunteer_header)
        self.volunteers_entry = QTextEdit()
        self.volunteers_entry.setFixedHeight(60)
        left.addWidget(self.volunteers_entry)

        operating_hours_label = QLabel('Operating hours')
        operating_hours_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        left.addWidget(operating_hours_label)
        self.operating_hours = QLineEdit('10:00 - 5:00')
        left.addWidget(self.operating_hours)

        lunch_frame = QFrame()
        lunch_frame.setFrameStyle(QFrame.Box | QFrame.Plain)
        lunch_row = QHBoxLayout(lunch_frame)

        timing_col = QVBoxLayout()
        lunch_label = QLabel('Lunch times?')
        lunch_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        timing_col.addWidget(lunch_label)
        self.lunch_timing_group = QButtonGroup(self)
        early_lunch = QRadioButton('Early')
        late_lunch = QRadioButton('Late')
        late_lunch.setChecked(True)
        self.lunch_timing_group.addButton(early_lunch, 0)
        self.lunch_timing_group.addButton(late_lunch, 1)
        timing_col.addWidget(early_lunch)
        timing_col.addWidget(late_lunch)
        lunch_row.addLayout(timing_col)

        hour_col = QVBoxLayout()
        hour_lunch_label = QLabel('Hour Lunches?')
        hour_lunch_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        hour_col.addWidget(hour_lunch_label)
        self.hour_lunch_group = QButtonGroup(self)
        hour_yes = QRadioButton('Yes')
        hour_no = QRadioButton('No')
        hour_yes.setChecked(True)
        self.hour_lunch_group.addButton(hour_yes, 1)
        self.hour_lunch_group.addButton(hour_no, 0)
        hour_col.addWidget(hour_yes)
        hour_col.addWidget(hour_no)
        lunch_row.addLayout(hour_col)

        left.addWidget(lunch_frame)

        create_btn = QPushButton('Create Blank')
        create_btn.setObjectName('primaryBtn')
        create_btn.clicked.connect(self.create_schedule)
        left.addWidget(create_btn)

        self.notes_text_box = QTextEdit()
        self.notes_text_box.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding
        )
        left.addWidget(self.notes_text_box, stretch=1)

        right = QVBoxLayout()
        main_layout.addLayout(right, stretch=2)

        self.schedule_area = QWidget()
        schedule_layout = QVBoxLayout(self.schedule_area)
        schedule_layout.setContentsMargins(0, 0, 0, 0)
        schedule_layout.addWidget(
            self.sheet_frame,
            0,
            Qt.AlignmentFlag.AlignTop)
        schedule_layout.addStretch(1)
        right.addWidget(self.schedule_area, stretch=3)

        self.inputs_host = QWidget()
        self.inputs_layout = QHBoxLayout(self.inputs_host)
        self.inputs_layout.addStretch()
        right.addWidget(self.inputs_host)


        bottom_row = QHBoxLayout()
        bottom_row.addStretch()
        open_btn = QPushButton('Open Schedule in Excel')
        open_btn.setObjectName('primaryBtn')
        open_btn.clicked.connect(self.open_excel)
        bottom_row.addWidget(open_btn)
        close_btn = QPushButton('Close MAS')
        close_btn.setObjectName('secondaryBtn')
        close_btn.clicked.connect(self._request_close)
        bottom_row.addWidget(close_btn)
        right.addLayout(bottom_row)

        self.load_notes()
        self.toast_manager = ToastManager(root)

    def _destroy_schedule_widgets(self) -> None:
        if self.inputs is not None:
            self.inputs.deleteLater()
            self.inputs = None

    def update_sheet(self):
        self.sheet_frame.update_sheet()
        if self.inputs is not None:
            self.inputs.standard_frame.update_coverage_indicators()

    def get_sheet_selection(self) -> dict | None:
        if not self.sheet_frame.has_schedule():
            return None
        regions = self.sheet_frame.table_view.selection_regions()
        if not regions:
            print('none selected')
            return None
        region = regions[0]
        time_start, time_end = region.time_range(self.df)
        workers = region.column_names(self.df)
        return {'workers': workers, 'time_start': time_start, 'time_end': time_end}

    def _primary_sheet_region(self) -> SelectionRegion | None:
        if not self.sheet_frame.has_schedule():
            return None
        regions = self.sheet_frame.table_view.selection_regions()
        if not regions:
            return None
        return regions[0]

    def copy_sheet_selection(self) -> None:
        region = self._primary_sheet_region()
        if region is None:
            return
        text = selection_values_to_tsv(region.values_2d(self.df))
        QApplication.clipboard().setText(text)

    def _paste_sheet_region(self, region: SelectionRegion, grid: list[list[str]]) -> None:
        apply_sheet_paste(self.df, region, grid)
        self.update_sheet()

    def paste_sheet_selection(self) -> None:
        region = self._primary_sheet_region()
        if region is None:
            return
        grid = parse_clipboard_tsv(QApplication.clipboard().text())
        if not grid:
            return
        self._perform_with_undo(self._paste_sheet_region, region, grid)

    def _clear_sheet_region(self, region: SelectionRegion) -> None:
        time_start, time_end = region.time_range(self.df)
        workers = region.column_names(self.df)
        self.df.loc[time_start:time_end, workers] = np.nan
        self.update_sheet()

    def cut_sheet_selection(self) -> None:
        region = self._primary_sheet_region()
        if region is None:
            return
        self.copy_sheet_selection()
        self._perform_with_undo(self._clear_sheet_region, region)

    def add_nonstandard_shift(self, selection):
        self.df.loc[
            selection['time_start'] : selection['time_end'], selection['workers']
        ] = selection['shift']
        self.nonstandard_shifts.append(selection)
        self.update_sheet()

    def add_standard_shift(self, shift):
        if not self.sheet_frame.has_schedule():
            return
        regions = self.sheet_frame.table_view.selection_regions()
        if not regions:
            workers = self.paid_workers + self.volunteers
            start, end = 0, len(self.df)
        else:
            region = regions[0]
            workers = region.column_names(self.df)
            start, end = region.from_row, region.upto_row

        failed_time_slots = []
        random.shuffle(workers)

        if SHIFT_INFO[shift]['isHour']:
            failed_time_slots = self._standard_full_hour_shift(shift, workers, start, end)
        if not SHIFT_INFO[shift]['isHour']:
            failed_time_slots = self._standard_half_hour_shift(shift, workers, start, end)

        if failed_time_slots and shift:
            msg = ', '.join(failed_time_slots)
            QMessageBox.warning(self, 'Warning', f'Failed to place {shift} at:\n{msg}')
        self.update_sheet()

    def _standard_half_hour_shift(self, shift, workers, start, end):
        failed_time_slots = []
        for curr_row in self.df.index[start:end]:
            if shift in self.df.loc[curr_row].values:
                continue
            nan_set = set(self.df.columns[self.df.loc[curr_row].isna()]) & set(workers)
            workers_with_nan = list(filter(lambda x: x in nan_set, workers))
            if workers_with_nan:
                worker_to_assign = next(w for w in workers if w in workers_with_nan)
                workers.remove(worker_to_assign)
                workers.append(worker_to_assign)
                self.df.at[curr_row, worker_to_assign] = shift
            if shift not in self.df.loc[curr_row].values:
                failed_time_slots.append(curr_row)
        return failed_time_slots

    def _standard_full_hour_shift(self, shift, workers, start, end):
        failed_time_slots = []
        for index, _row in self.df.iloc[start:end:2].iterrows():
            pos = self.df.index.get_loc(index)
            next_index = self.df.index[pos + 1]
            if shift in self.df.loc[index].values and shift in self.df.loc[next_index].values:
                continue
            workers_with_nan = (
                set(self.df.columns[self.df.iloc[pos].isna()])
                & set(self.df.columns[self.df.iloc[pos + 1].isna()])
                & set(workers)
            )
            if workers_with_nan:
                worker_to_assign = next(w for w in workers if w in workers_with_nan)
                workers.remove(worker_to_assign)
                workers.append(worker_to_assign)
                if shift not in self.df.loc[index].values:
                    self.df.at[index, worker_to_assign] = shift
                if shift not in self.df.loc[next_index].values:
                    self.df.at[next_index, worker_to_assign] = shift
            if shift not in self.df.loc[index].values:
                failed_time_slots.append(index)
            if shift not in self.df.loc[index].values:
                failed_time_slots.append(next_index)
        return failed_time_slots

    def _perform_with_undo(self, action_func, *args, **kwargs):
        state_before = self.df.copy()
        result = action_func(*args, **kwargs)
        if not self.df.equals(state_before):
            self.action_history_stack.append(state_before)
            self.action_redo_stack = []
            self.update_labels()
        else:
            print('No changes detected. No state saved')
        return result

    def undo(self):
        if self.action_history_stack:
            last_state = self.action_history_stack.pop()
            current_state = self.df.copy()
            self.action_redo_stack.append(current_state)
            self.df = last_state
            self.update_sheet()
        else:
            print('nothing left to undo.')
        self.update_labels()

    def redo(self):
        if self.action_redo_stack:
            next_state = self.action_redo_stack.pop()
            current_state = self.df.copy()
            self.action_history_stack.append(current_state)
            self.df = next_state
            self.update_sheet()
        else:
            print('nothing left to redo.')
        self.update_labels()

    def swap(self):
        if not self.sheet_frame.has_schedule():
            return
        regions = self.sheet_frame.table_view.selection_regions()
        if len(regions) < 2:
            print('insufficient selections.')
            QMessageBox.warning(self, 'Swap Error', 'Please select 2 equal size segments.')
            return
        sel1, sel2 = regions[0], regions[1]
        if sel1.row_count != sel2.row_count or sel1.col_count != sel2.col_count:
            print('incorrect size match')
            QMessageBox.warning(self, 'Swap Error', 'Incorrect size match.')
            return

        sel1_data = [np.nan if x == '' else x for x in sel1.values_flat(self.df)]
        sel2_data = [np.nan if x == '' else x for x in sel2.values_flat(self.df)]

        sel1_time_start, sel1_time_end = sel1.time_range(self.df)
        sel2_time_start, sel2_time_end = sel2.time_range(self.df)
        sel1_workers = sel1.column_names(self.df)
        sel2_workers = sel2.column_names(self.df)

        sel1_data_loc = self.df.loc[sel1_time_start:sel1_time_end, sel1_workers]
        sel2_data_loc = self.df.loc[sel2_time_start:sel2_time_end, sel2_workers]

        sel1_df = pd.DataFrame(sel1_data)
        sel1_df.index = sel2_data_loc.index
        sel1_df.columns = sel2_data_loc.columns

        sel2_df = pd.DataFrame(sel2_data)
        sel2_df.index = sel1_data_loc.index
        sel2_df.columns = sel1_data_loc.columns

        self.df.loc[sel2_time_start:sel2_time_end, sel2_workers] = sel1_df
        self.df.loc[sel1_time_start:sel1_time_end, sel1_workers] = sel2_df
        self.update_sheet()
        self.update_labels()

    def delete_column_clicked(self) -> None:
        removed_shifts = self._perform_with_undo(self.delete_column)
        if removed_shifts:
            shift_list = ', '.join(sorted(removed_shifts))
            self.toast_manager.show(
                f'standard shifts removed: {shift_list}',
                background_color=secondary_button_color,
                duration_ms=3000,
            )

    def delete_column(self) -> set[str] | None:
        if not self.sheet_frame.has_schedule():
            return None
        region = self._primary_sheet_region()
        if region is None:
            QMessageBox.warning(
                self, 'Delete Column', 'Please select a column to delete.'
            )
            return None
        workers = region.column_names(self.df)
        if not workers:
            return None

        standard_floor_shift_set = set(STANDARD_FLOOR_SHIFTS)
        removed_shifts: set[str] = set()
        for worker in workers:
            present = set(self.df[worker].dropna().unique())
            removed_shifts |= present & standard_floor_shift_set

        self.df = self.df.drop(columns=workers)
        for worker in workers:
            if worker in self.paid_workers:
                self.paid_workers.remove(worker)
            if worker in self.volunteers:
                self.volunteers.remove(worker)
        self.paid_workers_entry.setPlainText(', '.join(self.paid_workers))
        self.volunteers_entry.setPlainText(', '.join(self.volunteers))

        self.sheet_frame.table_view.set_regions([])
        self.update_sheet()
        return removed_shifts

    def show_balance_rules_dialog(self):
        if self.df.empty:
            QMessageBox.warning(
                self,
                'Balance Error',
                'No schedule created yet. Create a blank schedule first.',
            )
            return
        balance_rules_dialog = BalanceRulesDialog(self, on_apply=self._apply_balance_rules)
        balance_rules_dialog.show()

    def _apply_balance_rules(self, rules):
        self._perform_with_undo(lambda: self.auto_balance_shifts(rules))
        self.toast_manager.show('rules applied', background_color=primary_button_color)

    def auto_balance_shifts(self, balance_rules=None):
        balance_rules = balance_rules or default_balance_rules()
        max_iterations = 100

        for _iteration in range(max_iterations):
            found_violation = False
            for col_name, col_data in self.df.sample(frac=1,axis=1).items():
                col_series = self.df[col_name]
                for i in range(len(col_series)):
                    cell_value = col_series.iloc[i]
                    row_label = col_series.index[i]
                    is_nan = pd.isna(cell_value)
                    if not is_nan and cell_value not in SWAPPABLE_FLOOR_SHIFTS:
                        continue
                    col_violations_before = count_column_violations(col_series, balance_rules)
                    if col_violations_before == tuple(0 for _ in balance_rules):
                        continue
                    temp = col_series.copy()
                    temp.iloc[i] = np.nan
                    if count_column_violations(temp, balance_rules) >= col_violations_before:
                        continue
                    best_swap_col = None
                    best_score = total_violations(self.df, balance_rules)
                    for other_col_name, other_col_data in self.df.sample(frac=1,axis=1).items():
                        if other_col_name == col_name:
                            continue
                        candidate_value = self.df.at[row_label, other_col_name]
                        candidate_is_nan = pd.isna(candidate_value)
                        if not candidate_is_nan and candidate_value not in SWAPPABLE_FLOOR_SHIFTS:
                            continue
                        if direct_swap_blocked(balance_rules, cell_value, candidate_value):
                            continue
                        sim_df = self.df.copy()
                        sim_df.at[row_label, col_name] = candidate_value
                        sim_df.at[row_label, other_col_name] = cell_value
                        score = total_violations(sim_df, balance_rules)
                        if score < best_score:
                            best_score = score
                            best_swap_col = other_col_name
                    if best_swap_col is not None:
                        orig_value = self.df.at[row_label, col_name]
                        self.df.at[row_label, col_name] = self.df.at[row_label, best_swap_col]
                        self.df.at[row_label, best_swap_col] = orig_value
                        found_violation = True
                        break
                if found_violation:
                    break
            if not found_violation:
                break
        else:
            QMessageBox.warning(
                self,
                'Balance Warning',
                'Could not fully resolve all conflicts. '
                'Try running Balance again or adjust the schedule manually.',
            )
        self.update_sheet()

    def update_labels(self):
        if not self.inputs:
            return
        undo_len = len(self.action_history_stack)
        redo_len = len(self.action_redo_stack)
        self.inputs.undo_button.setText(f'Undo ({undo_len})')
        self.inputs.redo_button.setText(f'Redo ({redo_len})')

    def load_notes(self):
        try:
            with open('daily_notes.txt') as file:
                self.notes_text_box.setPlainText(file.read())
        except FileNotFoundError:
            self.notes_text_box.setPlainText(
                'Error: File not found.\nPlease make a file named daily_notes.txt and\n'
                'place it in the same folder as schedule.py'
            )
        except Exception as e:
            self.notes_text_box.setPlainText(f'An error occurred: {e}')

    def save_notes(self):
        with open('daily_notes.txt', 'w') as file:
            file.write(self.notes_text_box.toPlainText())

    def create_schedule(self):
        paid_workers_raw = self.paid_workers_entry.toPlainText()
        zachk_easter_egg_pattern = r"\b[Zz]ac[kh]?\b"
        paid_workers_text = re.sub(zachk_easter_egg_pattern, 'Zachk', paid_workers_raw) ################## << easter egg pattern
        self.paid_workers = paid_workers_text.split(', ')
        self.volunteers = self.volunteers_entry.toPlainText().split(', ')
        all_names = [name for name in self.paid_workers + self.volunteers if name]
        if len(all_names) != len(set(all_names)):
            self.toast_manager.show('duplicate names', background_color=secondary_button_color)
            return

        if self.sheet_frame.has_schedule():
            self.action_history_stack = []
            self.action_redo_stack = []
            self.update_labels()
            self._destroy_schedule_widgets()

        is_late_lunch = self.lunch_timing_group.checkedId()
        is_hour_lunch = self.hour_lunch_group.checkedId()

        start = 10
        end = 17
        hours_raw_text = self.operating_hours.text()
        pattern = r'(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})'
        match = re.search(pattern, hours_raw_text)
        if match:
            start = int(match.group(1))
            end = int(match.group(3)) + 12
            end_minutes = int(match.group(4))
            if end_minutes > 0:
                end += 1

        times = pd.to_datetime(
            [datetime.time(h, m).strftime('%H:%M') for h in range(start, end) for m in (0, 30)],
            format='%H:%M',
        ).strftime('%I:%M')
        if self.volunteers[0]:
            self.df = pd.DataFrame(columns=self.paid_workers + self.volunteers, index=times)
        else:
            self.df = pd.DataFrame(columns=self.paid_workers, index=times)

        self.fill_lunch(is_late_lunch, is_hour_lunch)
        if end >= 19:
            self.fill_dinner()

        height = 415 if end > 17 else 365
        self.sheet_frame.set_frame_size(975, height)
        self.sheet_frame.show_schedule()
        self.update_sheet()

        self.inputs = InputFrame(self)
        while self.inputs_layout.count():
            item = self.inputs_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.inputs_layout.addWidget(self.inputs)
        self.inputs.configure_action_buttons()
        self.inputs.standard_frame.update_coverage_indicators()

        self.update_labels()

    def fill_lunch(self, is_late_lunch, is_hour_lunch):
        possible_lunch_times = ['11:00', '11:30', '12:00', '12:30', '01:00', '01:30']
        if is_late_lunch:
            possible_lunch_times = ['01:00', '01:30', '12:00', '12:30', '11:00', '11:30']
        base_lunch_times = [time for time in possible_lunch_times if time in self.df.index]

        worker_groups = [self.paid_workers]
        if self.volunteers[0]:
            worker_groups.append(self.volunteers)

        for workers in worker_groups:
            workers = list(workers)
            random.shuffle(workers)
            lunch_times = list(base_lunch_times)
            for worker in workers:
                pos = self.df.index.get_loc(lunch_times[0])
                if is_hour_lunch:
                    self.df.at[self.df.index[pos], worker] = 'Lunch'
                    self.df.at[self.df.index[pos + 1], worker] = 'Lunch'
                    lunch_times.append(lunch_times.pop(0))
                    lunch_times.append(lunch_times.pop(0))
                else:
                    if lunch_times[0] == '11:00':
                        lunch_times.pop(0)
                        pos = self.df.index.get_loc(lunch_times[0])
                    self.df.at[self.df.index[pos], worker] = 'Lunch'
                    lunch_times.append(lunch_times.pop(0))

    def fill_dinner(self):
        possible_dinner_times = ['05:00', '05:30', '06:00', '06:30']
        base_dinner_times = [time for time in possible_dinner_times if time in self.df.index]

        worker_groups = [self.paid_workers]
        if self.volunteers[0]:
            worker_groups.append(self.volunteers)

        for workers in worker_groups:
            workers = list(workers)
            random.shuffle(workers)
            dinner_times = list(base_dinner_times)
            for worker in workers:
                pos = self.df.index.get_loc(dinner_times[0])
                # if dinner_times[0] == '05:00':
                #     dinner_times.pop(0)
                #     pos = self.df.index.get_loc(dinner_times[0])
                self.df.at[self.df.index[pos], worker] = 'Dinner'
                dinner_times.append(dinner_times.pop(0))

    def make_excel_file(self):
        self.df.index.name = f'{datetime.date.today().month}/{datetime.date.today().day}'
        with pd.ExcelWriter(RES_FILE_NAME, mode='w') as writer:
            self.df.to_excel(writer, sheet_name='Sheet1')
        wb = load_workbook(RES_FILE_NAME)
        ws = wb.active
        for cells_in_row in ws.iter_rows(min_row=2, max_col=len(self.df.columns) + 1):
            for cell in cells_in_row:
                cell_color = SHIFT_INFO.get(cell.internal_value)
                if cell_color:
                    raw = cell_color['color']
                    fg = raw[1:] if raw.startswith('#') else raw
                    cell.fill = PatternFill(patternType='solid', fgColor=fg)
        wb.save(RES_FILE_NAME)

    def open_excel(self):
        self.make_excel_file()
        try:
            self.open_file(RES_FILE_NAME)
        except FileNotFoundError:
            QMessageBox.warning(
                self, 'Excel Error', 'No schedule created yet, cannot open in Excel.'
            )
        except PermissionError:
            QMessageBox.warning(
                self,
                'Excel Error',
                'Please close the current excel sheet before opening the new one.',
            )
        except Exception as e:
            QMessageBox.warning(self, 'Excel Error', str(e))

    def open_file(self, filename: str) -> None:
        if sys.platform == 'win32':
            os.startfile(filename)
        else:
            opener = 'open' if sys.platform == 'darwin' else 'xdg-open'
            subprocess.call([opener, filename])

    def _request_close(self):
        self.close()

    def closeEvent(self, event: QCloseEvent):
        reply = QMessageBox.question(
            self,
            'Quit',
            'Are you sure you want to quit?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self.save_notes()
            event.accept()
        else:
            event.ignore()

def show_ui() -> None:
    app = QApplication(sys.argv)
    app.setFont(QFont(regular_font_family, regular_font_size))
    app.setStyleSheet(APP_STYLESHEET)
    window = ScheduleApp()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    show_ui()
