import copy
import datetime
import os
import random
import subprocess
import sys
import tkinter as tk
from tkinter import messagebox, ttk
from tksheet import Sheet
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# contains standard information about all shifts.
SHIFT_INFO = {
    'Trike': {'color':'#9999FF','isHour':False},
    'Gallery': {'color':'#5eb91e','isHour':False},
    'Back': {'color':'#E1EC24','isHour':False},
    'Front': {'color':'#9BC2E6','isHour':False},
    'Float': {'color':'#ffffff','isHour':False},
    'ENCA': {'color':'#FFD966','isHour':False},
    'Head': {'color':'#ffd221','isHour':False},
    'MOT': {'color':'#ffd221','isHour':False},
    'DIAR': {'color':'#b50934','isHour':False},
    'Lunch': {'color':'#7d7f7c','isHour':False},
    'Security': {'color':'#00b1d2','isHour':True},
    'Tickets': {'color':'#f6f9d4','isHour':True},
    'FLOAT': {'color':'#D3D3D3','isHour':False},
    'Badges': {'color':'#f6f9d4','isHour':True},
    'Project': {'color':'#f83dda','isHour':False},
    'GRGR': {'color':'#cf6498','isHour':True},
    'Manager': {'color':'#ACB9CA','isHour':True},
    'Securager': {'color':'#00b1d2','isHour':True},
    'STST': {'color':'#bf94e4','isHour':False},
    'MP': {'color':'#FFA500','isHour':False},
    'Museum Project': {'color':'3FFA500','isHour':True},
    'Training': {'color':'#FFD580','isHour':True},
    'Camp': {'color':'#c7ea46','isHour':True},
    'Retail': {'color':'#ffccd4','isHour':True},
    'Float 0': {'color':'#ffffff','isHour':False},
    'Float 1': {'color':'#90E4C1','isHour':False},
    'CORO': {'color':'#f4b183','isHour':False},
    'Pizza': {'color':"#f73939",'isHour':True},
    '': {'color':'#D3D3D3','isHour':False},
    None: {'color':'#D3D3D3','isHour':False}
}

# The selection of shifts that will show up in the listbox for adding Standard Shifts.
STANDARD_FLOOR_SHIFTS = ['Security', 'Trike', 'CORO', 'Gallery', 'Front', 'Back', 'Float 0', 'Float 1', 'ENCA']
SWAPPABLE_FLOOR_SHIFTS = ['Trike', 'CORO', 'Gallery', 'Front', 'Back', 'Float 0', 'Float 1', 'ENCA']

primary_button_color = "#EDD863"
primary_button_hover_color = "#E1D591"
secondary_button_color = "#6A2E35"
secondary_button_hover_color = "#78454C"
# redo_button_color = "#E0ACD5"
# redo_hover_color = "#E6C7DF"
# generic_button_color = "#5C6B73"
# generic_hover_color = "#687278"


if sys.platform == 'win32':
    RES_FILE_NAME = f'momath_schedule_{datetime.date.today()}.xlsx'
else:
    RES_FILE_NAME = f'/tmp/momath_schedule_{datetime.date.today()}.xlsx'


# ---------------------------------------------------------------------------
# Auto-balance rules (each rule is an object; toggle .enabled in auto_balance_shifts)
# ---------------------------------------------------------------------------

BEFORE_1PM_CUTOFF = '01:00'


class ShiftBalanceRule:
    """One schedulable constraint for auto_balance_shifts."""

    name = 'base'
    description = ''

    def __init__(self, enabled=True):
        self.enabled = enabled

    def count(self, col_series):
        if not self.enabled:
            return 0
        return self._count(col_series)

    def _count(self, col_series):
        raise NotImplementedError


class NoDuplicateConsecutiveRule(ShiftBalanceRule):
    """No duplicate consecutive swappable floor shifts in a column."""

    name = 'no_duplicate_consecutive'
    description = 'No duplicate consecutive shifts.'

    def _count(self, col_series):
        values = col_series.tolist()
        violations = 0
        for i in range(len(values) - 1):
            val, next_val = values[i], values[i + 1]
            if val in SWAPPABLE_FLOOR_SHIFTS and val == next_val:
                violations += 1
        return violations


class NoTrikeCoroAdjacencyRule(ShiftBalanceRule):
    """Trike and CORO must not appear in adjacent rows (either order)."""

    name = 'no_trike_coro_adjacency'
    description = 'No consecutiveTrike/CORO.'

    def _count(self, col_series):
        values = col_series.tolist()
        violations = 0
        for i in range(len(values) - 1):
            val, next_val = values[i], values[i + 1]
            if {val, next_val} == {'Trike', 'CORO'}:
                violations += 1
        return violations


class MaxOneTrikeCoroBefore1pmRule(ShiftBalanceRule):
    """At most one Trike or CORO in rows before 1 PM ('01:00')."""

    name = 'max_one_trike_coro_before_1pm'
    description = "Minimize Trike/CORO before 1 PM."

    def _count(self, col_series):
        index = col_series.index.tolist()
        try:
            cutoff_pos = index.index(BEFORE_1PM_CUTOFF)
        except ValueError:
            cutoff_pos = len(index)

        trike_coro_before_1pm = sum(
            1 for i, val in enumerate(col_series.tolist())
            if i < cutoff_pos and val in ('Trike', 'CORO')
        )
        if trike_coro_before_1pm > 1:
            return trike_coro_before_1pm - 1
        return 0


class NoTrikeAdjacentLunchRule(ShiftBalanceRule):
    """Trike must not be directly adjacent to Lunch (either order)."""

    name = 'no_trike_adjacent_lunch'
    description = 'Trike not adjacent to Lunch.'

    def _count(self, col_series):
        values = col_series.tolist()
        violations = 0
        for i in range(len(values) - 1):
            if {values[i], values[i + 1]} == {'Trike', 'Lunch'}:
                violations += 1
        return violations


def default_balance_rules():
    """Rules in priority order (index 0 = highest priority)."""
    return [
        NoDuplicateConsecutiveRule(),
        NoTrikeCoroAdjacencyRule(),
        MaxOneTrikeCoroBefore1pmRule(),
        NoTrikeAdjacentLunchRule(),
    ]


def set_balance_rule_enabled(rules, name, enabled):
    """Enable or disable a rule by its .name (see rule classes above)."""
    for rule in rules:
        if rule.name == name:
            rule.enabled = enabled
            return
    raise ValueError(f'Unknown balance rule: {name!r}')


def count_column_violations(col_series, rules):
    """
    Count violations per rule. Returns a priority tuple; tuple comparison
    enforces rule order (fewer high-priority violations always wins).
    """
    return tuple(rule.count(col_series) for rule in rules)


def total_violations(df, rules):
    totals = [0] * len(rules)
    for col in df.columns:
        for bucket, count in enumerate(count_column_violations(df[col], rules)):
            totals[bucket] += count
    return tuple(totals)


def introduces_no_new_violations(df, row_label, col_name, new_value, rules):
    """
    Per-column helper: simulate placing `new_value` into df.at[row_label, col_name].
    Returns True if the column priority tuple does not get worse.

    Not used as a swap gate in auto_balance_shifts; swaps are accepted only when
    total_violations improves schedule-wide (lexicographic rule order).
    df is not modified.
    """
    current_score = count_column_violations(df[col_name], rules)

    df_copy = df.copy()
    df_copy.at[row_label, col_name] = new_value
    new_score = count_column_violations(df_copy[col_name], rules)

    return new_score <= current_score


def format_balance_rule_line(index, rule):
    """Display text for one rule in the balance dialog listbox."""
    flag = '[on]' if rule.enabled else '[off]'
    return f'{flag}  {index + 1}.  {rule.description}'


def _bind_button_hover(button, normal_color, hover_color):
    button.configure(background=normal_color)
    button.bind('<Enter>', lambda _e: button.configure(background=hover_color))
    button.bind('<Leave>', lambda _e: button.configure(background=normal_color))


class BalanceRulesDialog(tk.Toplevel):
    """Modal dialog to enable, disable, and reorder balance rules before applying."""

    def __init__(self, parent, on_apply):
        super().__init__(parent)
        self.on_apply = on_apply
        self.title('Balance Shifts')
        self.configure(background='lightblue')
        self.transient(parent)
        self.grab_set()

        self.rules = copy.deepcopy(default_balance_rules())

        intro = tk.Label(
            self,
            text='Rules run top to bottom. Order matters.',
            background='lightblue',
            justify='left',
            wraplength=520,
        )
        intro.pack(anchor='w', padx=10, pady=(10, 5))

        content = tk.Frame(self, relief=tk.GROOVE, borderwidth=2, background='lightblue')
        content.pack(fill='both', expand=True, padx=10, pady=5)

        self.rule_listbox = tk.Listbox(content, width=72, height=6, selectmode='single')
        self.rule_listbox.pack(side='left', fill='both', expand=True, padx=(5, 0), pady=5)

        controls = tk.Frame(content, background='lightblue')
        controls.pack(side='left', fill='y', padx=5, pady=5)

        toggle_btn = tk.Button(controls, text='Enable / Disable', command=self._toggle_selected, width=14)
        toggle_btn.pack(pady=(0, 5))

        up_btn = tk.Button(controls, text='Move Up', command=self._move_up, width=14)
        up_btn.pack(pady=5)

        down_btn = tk.Button(controls, text='Move Down', command=self._move_down, width=14)
        down_btn.pack(pady=5)

        action_row = tk.Frame(self, background='lightblue')
        action_row.pack(fill='x', padx=10, pady=(5, 10))

        cancel_btn = tk.Button(
            action_row,
            text='Cancel',
            command=self.destroy,
            width=12,
            foreground='#ffffff',
        )
        cancel_btn.pack(side='right', padx=(5, 0))
        _bind_button_hover(cancel_btn, secondary_button_color, secondary_button_hover_color)

        apply_btn = tk.Button(action_row, text='Apply', command=self._apply, width=12)
        apply_btn.pack(side='right')
        _bind_button_hover(apply_btn, primary_button_color, primary_button_hover_color)

        self.refresh_rule_list()
        self.rule_listbox.selection_set(0)

    def refresh_rule_list(self):
        selection = self.rule_listbox.curselection()
        selected_index = selection[0] if selection else 0

        self.rule_listbox.delete(0, tk.END)
        for index, rule in enumerate(self.rules):
            self.rule_listbox.insert(tk.END, format_balance_rule_line(index, rule))

        if self.rules:
            selected_index = min(selected_index, len(self.rules) - 1)
            self.rule_listbox.selection_set(selected_index)
            self.rule_listbox.see(selected_index)

    def _selected_index(self):
        selection = self.rule_listbox.curselection()
        if not selection:
            return None
        return selection[0]

    def _toggle_selected(self):
        index = self._selected_index()
        if index is None:
            return
        self.rules[index].enabled = not self.rules[index].enabled
        self.refresh_rule_list()

    def _move_up(self):
        index = self._selected_index()
        if index is None or index == 0:
            return
        self.rules[index], self.rules[index - 1] = self.rules[index - 1], self.rules[index]
        self.refresh_rule_list()
        self.rule_listbox.selection_set(index - 1)
        self.rule_listbox.see(index - 1)

    def _move_down(self):
        index = self._selected_index()
        if index is None or index >= len(self.rules) - 1:
            return
        self.rules[index], self.rules[index + 1] = self.rules[index + 1], self.rules[index]
        self.refresh_rule_list()
        self.rule_listbox.selection_set(index + 1)
        self.rule_listbox.see(index + 1)

    def _apply(self):
        self.on_apply(self.rules)
        self.destroy()


# ---------------------------------------------------------------------------


class ScheduleApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('MoMath Automatic Scheduler')
        self.geometry('1500x800')
        self.configure(background='lightblue')

        self.nonstandard_shifts = []
        self.action_history_stack = []
        self.action_redo_stack = []
        self.df = pd.DataFrame()
        self.sheet_frame = None
        self.create_widgets()

    def create_widgets(self):
        """Initialize all app widgets."""
        self.create_worker_inputs()
        self.create_radio_buttons()
        self.create_action_buttons()
        self.create_notes_box()

    def create_worker_inputs(self):
        """Initialize the worker entry boxes."""
        # Paid workers input
        tk.Label(self, text='Paid workers (comma separated)').pack(anchor='w', pady=10, padx=10)
        self.paid_workers_entry = tk.Text(self, width=40, height=3, wrap='word')
        #self.paid_workers_entry.insert(0, get_ft()) # get_ft() not added yet.
        #self.paid_workers_entry.insert(0, "placeholder")
        self.paid_workers_entry.pack(anchor='w', pady=5, padx=10)

        # Volunteers input
        tk.Label(self, text='Volunteers (comma separated)').pack(anchor='w', pady=5, padx=10)
        self.volunteers_entry = tk.Text(self, width=40, height=3, wrap='word')
        self.volunteers_entry.pack(anchor='w', pady=5, padx=10)

    def create_radio_buttons(self):
        """Initialize radio buttons."""
        tk.Label(self, text='Operating hours').pack(anchor='w', pady=5, padx=10)
        self.operating_hours = tk.Entry(self, width=20)
        self.operating_hours.insert(0,"10:00 - 5:00")
        self.operating_hours.pack(anchor='w', pady=5, padx=10)

        self.radio0 = tk.IntVar(value=1)
        tk.Label(self, text='Early or late lunches today?').pack(anchor='w', pady=5, padx=10)
        tk.Radiobutton(self, text='Early', variable=self.radio0, value=0).pack(anchor='w', pady=5, padx=10)
        tk.Radiobutton(self, text='Late', variable=self.radio0, value=1).pack(anchor='w', pady=5, padx=10)

        self.radio1 = tk.IntVar(value=1)
        tk.Label(self, text='Hour Lunches?').pack(anchor='w', pady=5, padx=10)
        tk.Radiobutton(self, text='Yes', variable=self.radio1, value=1).pack(anchor='w', pady=5, padx=10)
        tk.Radiobutton(self, text='No', variable=self.radio1, value=0).pack(anchor='w', pady=5, padx=10)


    def create_action_buttons(self):
        """Initialize main buttons."""
        create_schedule_button = tk.Button(self, text="Create Blank", command=self.create_schedule)
        create_schedule_button.pack(anchor='w', pady=20, padx=10)

        open_schedule_button = tk.Button(self, text='Open Schedule in Excel', command=self.open_excel, height=2, width=20)
        open_schedule_button.place(relx=0.993, rely=0.89, anchor='e')

        close_button = tk.Button(self, text='Close MAS', command=self.close, height=2, width=10)
        close_button.place(relx=0.993, rely=0.95, anchor='e')

    def create_notes_box(self):
        """Initialize Notes textbox."""
        self.notes_text_box = tk.Text(self, wrap='word', width=60, height=15)
        self.notes_text_box.place(relx=0.993, rely=0.70, anchor='e')
        self.load_notes()

    def destroy_sheet(self) -> None:
        """Destroys sheets."""
        for child in self.winfo_children():
            if isinstance(child, Sheet):
                child.destroy()

    def update_sheet(self):
        """Update the display sheet with the information in the dataframe."""
        self.sheet_frame.update_sheet()
        return
    
    def destroy_frame(window: tk.Tk) -> None:
        """Destroys frame."""
        for child in window.winfo_children():
            if isinstance(child, tk.Frame):
                child.destroy()

    def get_sheet_selection(self) -> dict:
        """Gets the selection of the sheet."""
        if not self.sheet_frame.sheet.get_all_selection_boxes():
            print("none selected")
            return
        selection = self.sheet_frame.sheet.get_all_selection_boxes()[0] # get first selection only
        time_start = self.df.iloc[selection.from_r].name
        time_end = self.df.iloc[selection.upto_r-1].name

        workers = self.sheet_frame.sheet.headers()[selection.from_c:selection.upto_c]

        return {"workers": workers, "time_start": time_start, "time_end": time_end}

    def add_nonstandard_shift(self, selection):
        """Handle adding nonstandard shifts to the dataframe and the display."""
        self.df.loc[selection["time_start"]:selection["time_end"], selection["workers"]] = selection["shift"]
        self.nonstandard_shifts.append(selection)
        self.update_sheet()

    
    def add_standard_shift(self, shift):
        """Handle adding standard shifts to the dataframe and the display."""
        # Determine workers to use
        if not self.sheet_frame.sheet.get_all_selection_boxes():
            selection = None
            workers = self.paid_workers + self.volunteers
        else:
            selection = self.sheet_frame.sheet.get_all_selection_boxes()[0]
            workers = self.sheet_frame.sheet.headers()[selection.from_c:selection.upto_c]
        
        failed_time_slots = []
        random.shuffle(workers)

        ###
        # add in a section to only do the rows selected.
        if selection == None:
            start = 0
            end = len(self.df)
        else:
            start = selection[0]
            end = selection[2]
        # add check to get the hours selection.
        # Be careful about hour selection
        ###

        # if hour shift -> 
        if SHIFT_INFO[shift]['isHour']:
            failed_time_slots = self._standard_full_hour_shift(shift, workers, start, end)

        # if half-hour shift ->
        if not SHIFT_INFO[shift]['isHour']:
            failed_time_slots = self._standard_half_hour_shift(shift, workers, start, end)
    
        
        # Show warning if any slots failed
        if failed_time_slots and shift:
            msg = ', '.join(failed_time_slots)
            messagebox.showwarning('Warning', f'Failed to place {shift} at:\n{msg}')
        
        self.update_sheet()
    
    def _standard_half_hour_shift(self, shift, workers, start, end):
        """
        Internal function to add 1 copy of shift at each half hour to the given set of workers.
        
        :param shift: a string. 
        :param workers: list of columns in dataframe to add shift to.
        """
        failed_time_slots = []

        for curr_row in self.df.index[start:end]:
            if shift in self.df.loc[curr_row].values:
                # only apply a single copy of a given shift per schedule.
                continue
            nan_set = set(self.df.columns[self.df.loc[curr_row].isna()]) & set(workers)
            workers_with_nan = list(filter(lambda x: x in nan_set, workers))
            
            if workers_with_nan:
                worker_to_assign = next(w for w in workers if w in workers_with_nan)
                workers.remove(worker_to_assign)
                workers.append(worker_to_assign)
                self.df.at[curr_row, worker_to_assign] = shift
            
            if shift not in self.df.loc[curr_row].values:
                failed_time_slots.append(curr_row)

        return failed_time_slots

    def _standard_full_hour_shift(self, shift, workers, start, end):
        """
        Internal function to add 1 copy of shift at each hour to the given set of workers.
        
        :param shift: a string.
        :param workers: list of columns in dataframe to add shift to.
        """
        failed_time_slots = []

        for index, row in self.df.iloc[start:end:2].iterrows():
            pos = self.df.index.get_loc(index)
            next_index = self.df.index[pos+1]

            if shift in self.df.loc[index].values and shift in self.df.loc[next_index].values:
                # only apply a single copy of a given shift per schedule.
                continue

            workers_with_nan = set(self.df.columns[self.df.iloc[pos].isna()]) & set(self.df.columns[self.df.iloc[pos+1].isna()]) & set(workers)
            
            if workers_with_nan:
                worker_to_assign = next(w for w in workers if w in workers_with_nan)
                workers.remove(worker_to_assign)
                workers.append(worker_to_assign)
                if shift not in self.df.loc[index].values:
                    self.df.at[index, worker_to_assign] = shift
                if shift not in self.df.loc[next_index].values:
                    self.df.at[next_index, worker_to_assign] = shift
            
            if shift not in self.df.loc[index].values:
                failed_time_slots.append(index)
            if shift not in self.df.loc[index].values:
                failed_time_slots.append(next_index)

        return failed_time_slots


    def _perform_with_undo(self, action_func, *args, **kwargs):
        """Execute an action and only save state if changes were made."""
        state_before = self.df.copy()
        
        # Perform the action
        result = action_func(*args, **kwargs)
        
        # Check if anything changed (old save_state)
        if not self.df.equals(state_before):
            self.action_history_stack.append(state_before)
            self.action_redo_stack = []
            self.update_labels()
        else:
            print("No changes detected. No state saved")
        
        return result
 
    def undo(self):
        if self.action_history_stack:
            last_state = self.action_history_stack.pop()
            current_state = self.df.copy()
            self.action_redo_stack.append(current_state)
            self.df = last_state
            self.update_sheet()
        else:
            print("nothing left to undo.") 
        self.update_labels() 
    
    def redo(self):
        if self.action_redo_stack:
            next_state = self.action_redo_stack.pop()
            current_state = self.df.copy()
            self.action_history_stack.append(current_state)
            self.df = next_state
            self.update_sheet()
        else:
            print("nothing left to redo.")
        self.update_labels()

    def swap(self):
        """Swaps a pair of selections if they are equal size. Action performed with undo."""
        # get first and second selections. 
        if len(self.sheet_frame.sheet.get_all_selection_boxes()) < 2:
            print("insufficient selections.")
            messagebox.showwarning('Swap Error', 'Please select 2 equal size segments.')
            return
        sel1 = self.sheet_frame.sheet.get_all_selection_boxes()[0]
        sel1_x = sel1[3]-sel1[1]
        sel1_y = sel1[2]-sel1[0]
        sel2 = self.sheet_frame.sheet.get_all_selection_boxes()[1]
        sel2_x = sel2[3]-sel2[1]
        sel2_y = sel2[2]-sel2[0]

        if sel1_x != sel2_x or sel1_y != sel2_y:
            print("incorrect size match")
            messagebox.showwarning('Swap Error', 'Incorrect size match.')
            return
        
        if sel1_x == 1 and sel1_y == 1:
            sel1_data = [self.sheet_frame.sheet.get_data(*sel1)]
            sel2_data = [self.sheet_frame.sheet.get_data(*sel2)]
        else:
            sel1_data = self.sheet_frame.sheet.get_data(*sel1)
            sel2_data = self.sheet_frame.sheet.get_data(*sel2)
        
        sel1_data = [np.nan if x == '' else x for x in sel1_data]
        sel2_data = [np.nan if x == '' else x for x in sel2_data]
        
        sel1_time_start = self.df.iloc[sel1.from_r].name
        sel1_time_end = self.df.iloc[sel1.upto_r-1].name
        sel1_workers = self.sheet_frame.sheet.headers()[sel1.from_c:sel1.upto_c]

        sel2_time_start = self.df.iloc[sel2.from_r].name
        sel2_time_end = self.df.iloc[sel2.upto_r-1].name
        sel2_workers = self.sheet_frame.sheet.headers()[sel2.from_c:sel2.upto_c]

        sel1_data_loc = self.df.loc[sel1_time_start:sel1_time_end, sel1_workers]
        sel2_data_loc = self.df.loc[sel2_time_start:sel2_time_end, sel2_workers]

        sel1_df = pd.DataFrame(sel1_data)
        sel1_df.index = sel2_data_loc.index
        sel1_df.columns = sel2_data_loc.columns

        sel2_df = pd.DataFrame(sel2_data)
        sel2_df.index = sel1_data_loc.index
        sel2_df.columns = sel1_data_loc.columns

        # assignment.
        self.df.loc[sel2_time_start:sel2_time_end, sel2_workers] = sel1_df
        self.df.loc[sel1_time_start:sel1_time_end, sel1_workers] = sel2_df

        # update labels.
        self.update_sheet()
        self.update_labels()

    def show_balance_rules_dialog(self):
        """Open the balance rules dialog; apply runs auto_balance with undo."""
        if self.df.empty:
            messagebox.showwarning('Balance Error', 'No schedule created yet. Create a blank schedule first.')
            return
        BalanceRulesDialog(self, on_apply=self._apply_balance_rules)

    def _apply_balance_rules(self, rules):
        self._perform_with_undo(lambda: self.auto_balance_shifts(rules))

    def auto_balance_shifts(self, balance_rules=None):
        """
        Iteratively resolve shift violations by swapping cells within the same row.
        Only SWAPPABLE_FLOOR_SHIFTS and NaN may be moved.

        Rule priority and enabled flags come from balance_rules (list order).
        """
        balance_rules = balance_rules or default_balance_rules()

        max_iterations = 100

        for iteration in range(max_iterations):
            found_violation = False

            for col in self.df.columns:
                col_series = self.df[col]

                for i in range(len(col_series)):
                    cell_value = col_series.iloc[i]
                    row_label = col_series.index[i]

                    is_nan = pd.isna(cell_value)
                    if not is_nan and cell_value not in SWAPPABLE_FLOOR_SHIFTS:
                        continue

                    col_violations_before = count_column_violations(col_series, balance_rules)
                    if col_violations_before == tuple(0 for _ in balance_rules):
                        continue

                    temp = col_series.copy()
                    temp.iloc[i] = np.nan
                    if count_column_violations(temp, balance_rules) >= col_violations_before:
                        continue

                    best_swap_col = None
                    best_score = total_violations(self.df, balance_rules)

                    for other_col in self.df.columns:
                        if other_col == col:
                            continue

                        candidate_value = self.df.at[row_label, other_col]

                        candidate_is_nan = pd.isna(candidate_value)
                        if not candidate_is_nan and candidate_value not in SWAPPABLE_FLOOR_SHIFTS:
                            continue

                        sim_df = self.df.copy()
                        sim_df.at[row_label, col] = candidate_value
                        sim_df.at[row_label, other_col] = cell_value
                        score = total_violations(sim_df, balance_rules)

                        if score < best_score:
                            best_score = score
                            best_swap_col = other_col

                    if best_swap_col is not None:
                        orig_value = self.df.at[row_label, col]
                        self.df.at[row_label, col] = self.df.at[row_label, best_swap_col]
                        self.df.at[row_label, best_swap_col] = orig_value
                        found_violation = True
                        break

                if found_violation:
                    break

            if not found_violation:
                break
        else:
            messagebox.showwarning(
                'Balance Warning',
                'Could not fully resolve all conflicts. '
                'Try running Balance again or adjust the schedule manually.'
            )

        self.update_sheet()

    def update_labels(self):
        undo_len = len(self.action_history_stack)
        undo_label = f"Undo ({undo_len})"
        redo_len = len(self.action_redo_stack)
        redo_label = f"Redo ({redo_len})"
        self.inputs.undo_button.configure(text=undo_label)
        self.inputs.redo_button.configure(text=redo_label)

    def load_notes(self):
        try:
            with open('daily_notes.txt') as file:
                content = file.read()
                self.notes_text_box.delete('1.0', tk.END)
                self.notes_text_box.insert(tk.END, content)
        except FileNotFoundError:
            self.notes_text_box.delete('1.0', tk.END)
            self.notes_text_box.insert(tk.END, 'Error: File not found.\nPlease make a file named daily_notes.txt and\nplace it in the same folder as schedule.py')
        except Exception as e:
            self.notes_text_box.delete('1.0', tk.END)
            self.notes_text_box.insert(tk.END, f'An error occurred: {e}')

    def save_notes(self):
        file_content = self.notes_text_box.get('1.0', tk.END)
        with open('daily_notes.txt', 'w') as file:
            file.write(file_content)

    def create_schedule(self):
        if self.sheet_frame:
            # clear previous UI element if button is clicked more than once.
            self.action_history_stack = []
            self.action_redo_stack = []
            self.update_labels()
            self.destroy_frame()


        self.paid_workers = self.paid_workers_entry.get("1.0","end-1c").split(', ')
        self.volunteers = self.volunteers_entry.get("1.0","end-1c").split(', ')
        is_late_lunch = self.radio0.get() # early or late lunch (0 or 1)
        is_hour_lunch = self.radio1.get() 

        start = 10
        end = 17

        # using regex to get integer from input museum operating hours.
        hours_raw_text = self.operating_hours.get()
        pattern = r'(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})'
        match = re.search(pattern, hours_raw_text)

        if match:
            start = int(match.group(1))
            end = int(match.group(3)) + 12
            end_minutes = int(match.group(4))
            if end_minutes > 0:
                end += 1

        times = pd.to_datetime([datetime.time(h, m).strftime('%H:%M') for h in range(start, end) for m in (0, 30)], format='%H:%M').strftime('%I:%M')
        if self.volunteers[0]: # if no volunteers entered, entrybox still returns empty string. So list = ['']
            self.df = pd.DataFrame(columns=self.paid_workers + self.volunteers, index=times)
        else:
            self.df = pd.DataFrame(columns=self.paid_workers, index=times)

        self.fill_lunch(is_late_lunch, is_hour_lunch)

        self.sheet_frame = sheetFrame(controller = self)
        self.sheet_frame.pack_propagate(False)


        if end > 17: # more height to accomidate longer schedule
            self.sheet_frame.place(height = 415, width = 975, relx=.992, rely=.0125, anchor='ne')
        else:
            self.sheet_frame.place(height = 365, width = 975, relx=.992, rely=.0125, anchor='ne')

        self.inputs = inputFrame(controller = self)
        self.update_sheet()
    
    def fill_lunch(self, is_late_lunch, is_hour_lunch):
        '''
        Fill the dataframe with lunches upon initialization of a blank sheet.
        
        :param self: app controller
        :param is_late_lunch: True if late lunches requested, False if early lunches requested
        :param is_hour_lunch: True for 1-hour lunch blocks, False for 30-minute lunch blocks
        '''
        # Define all 30-minute lunch time slots
        possible_lunch_times = ["11:00", "11:30", "12:00", "12:30", "01:00", "01:30"]
        
        # Get all workers (paid + volunteers if applicable)
        workers = self.paid_workers + self.volunteers if self.volunteers[0] else self.paid_workers
        random.shuffle(workers)
        
        # Reverse order for late lunches
        if is_late_lunch:
            possible_lunch_times = ["01:00", "01:30", "12:00", "12:30", "11:00", "11:30"]
            #lunch_times.reverse()
        
        lunch_times = [time for time in possible_lunch_times if time in self.df.index]
        
        # Assign lunches to workers
        for worker in workers:

            pos = self.df.index.get_loc(lunch_times[0])
            
            if is_hour_lunch:
                # Assign 1-hour lunch (two consecutive 30-min blocks)
                self.df.at[self.df.index[pos], worker] = 'Lunch'
                self.df.at[self.df.index[pos + 1], worker] = 'Lunch'
                # Rotate lunch times twice to move to next hour slot
                lunch_times.append(lunch_times.pop(0))
                lunch_times.append(lunch_times.pop(0))
            else:
                if lunch_times[0] == '11:00':
                    lunch_times.pop(0)
                    pos = self.df.index.get_loc(lunch_times[0])
                # Assign 30-minute lunch (alternating first/second half)
                self.df.at[self.df.index[pos], worker] = 'Lunch'
                # Rotate lunch times once
                lunch_times.append(lunch_times.pop(0))
    
    def make_excel_file(self):
        """Converts dataframe into excel file."""

        self.df.index.name = f'{datetime.date.today().month}/{datetime.date.today().day}'

        with pd.ExcelWriter(RES_FILE_NAME, mode='w') as writer:
            self.df.to_excel(writer, sheet_name='Sheet1')

        wb = load_workbook(RES_FILE_NAME)
        ws = wb.active

        for cells_in_row in ws.iter_rows(min_row=2,max_col=len(self.df.columns)+1):  # colors for excel
            for cell in cells_in_row:
                cell_color = SHIFT_INFO.get(cell.internal_value)
                if cell_color:
                    cell.fill = PatternFill(patternType='solid', fgColor=cell_color['color'][1:])

        wb.save(RES_FILE_NAME)
        return
        

    def open_excel(self):
        self.make_excel_file()
        try:
            self.open_file(RES_FILE_NAME)
        except FileNotFoundError:
            messagebox.showwarning('Excel Error', 'No schedule created yet, cannot open in Excel.')
        except PermissionError:
            messagebox.showwarning('Excel Error', 'Please close the current excel sheet before opening the new one.')
        except Exception as e:
            messagebox.showwarning('Excel Error', e)
    
    def open_file(self, filename: str) -> None:
        """Opens file depending on OS."""
        if sys.platform == 'win32':
            os.startfile(filename)
        else:
            opener = 'open' if sys.platform == 'darwin' else 'xdg-open'
            subprocess.call([opener, filename])

    def close(self):
        if messagebox.askyesno("Quit", "Are you sure you want to quit?"):
            self.save_notes()
            self.destroy()
    

class sheetFrame(tk.Frame):
    def __init__(self, controller: ScheduleApp):
        super().__init__(controller)
        self.controller = controller
        self.sheet = None

    def create_sheet(self, output_df):
        self.sheet = Sheet(self,
                        data=output_df.values.tolist(),
                        headers=output_df.columns.tolist(),
                        row_index=output_df.index.tolist(),
                        auto_resize_columns=True,
                        auto_resize_rows=True,
                        empty_horizontal=True,
                        empty_vertical=True
                        )
        self.sheet.enable_bindings("ctrl_select", "drag_select","single_select","column_select")
        self.sheet.disable_bindings("column_width_resize", "row_height_resize", "move_columns", "move_rows", "column_height_resize", "row_width_resize", "rc_menu")
        self.sheet.pack(fill="both", side='right',expand=True)
        # column_list = [60] * (len(output_df.columns)-1)
        # column_list.append(20)
        # self.sheet.set_column_widths(column_list)
        self.sheet.readonly_columns(columns=[i for i, _ in enumerate(output_df.columns)], readonly=True)

        # set column width here based on width of sheetFrame. 975 pixels.
        # total_width = 975
        # num_cols = len(output_df.columns)
        # initial_col_width = total_width/num_cols
        # free_slot_width = initial_col_width*0.25
        # available_width = total_width - free_slot_width
        # standard_col_width = available_width/(num_cols-1)

        # col_width_list = [standard_col_width] * (num_cols-1)
        # col_width_list.append(free_slot_width)
        # self.sheet.set_column_widths(col_width_list)


    def update_sheet(self):
        """Update the display sheet with the information in the dataframe."""
        df = self.controller.df.copy()
        #df['Free Slots'] = df.isnull().sum(axis=1)

        # if self.sheet:
        #     print("DEBUG:: update_sheet(): column widths")
        #     print(self.sheet.column_width(column=0  ))

        try:
            if not self.sheet:
                # This method of if else allows the sheet to not flicker upon every update.
                    # except now it is flickering because of the column width setup.
                # The fillna('') is only to display the gray cells instead of nan.
                self.create_sheet(df.fillna(""))
            else:
                self.sheet.set_sheet_data(data=df.fillna("").values.tolist())

            self.color_format()
            # column_list = [60] * (len(df.columns)-1)
            # column_list.append(20)
            # self.sheet.set_column_widths(column_list)
        
            # print("DEBUG:: update_sheet(): column widths")
            # print(self.sheet.column_width(column=0))

        except tk.TclError:
            pass

    def destroy_sheet(self):
        """Destroys sheets."""
        for child in self.winfo_children():
            if isinstance(child, Sheet):
                child.destroy()

    def color_format(self):
        """Apply color formatting to sheet."""
        self.sheet.dehighlight_all()

        for row_num, row in enumerate(self.sheet):
            #last_column = len(row)-1
            for col_num, shift in enumerate(row):
                #if col_num == last_column:
                    # # apply color gradient to the 'free slots' column. Each cell is the number of empty slots in the row.
                    # color_gradient_index = shift
                    # if color_gradient_index > 5:
                    #     color_gradient_index = 5
                    # color_gradient = ["#a41900","#db2100","#ff4827","#ff7962","#ffb6a9","#ffdbd4"]
                    # self.sheet.highlight_cells(row=row_num, column=col_num, bg=color_gradient[color_gradient_index], redraw=False)
                    # continue
                if pd.isna(shift):
                    self.sheet.highlight_cells(row=row_num, column=col_num, bg=None, redraw=False)
                elif shift in SHIFT_INFO:
                    self.sheet.highlight_cells(row=row_num, column=col_num, bg=SHIFT_INFO[shift]['color'], redraw=False)
                else:
                    self.sheet.highlight_cells(row=row_num, column=col_num, bg=None, redraw=False)
        self.sheet.refresh()

class inputFrame(tk.Frame):
    def __init__(self, controller: ScheduleApp):
        super().__init__(controller)
        self.controller = controller
        self.configure(background='lightblue')
        self.place(relx= .5 + 240/1500, y = 438, anchor='ne')
        self.create_widgets()

    def create_widgets(self):
        """
        Create and place every button and interactable element used during schedule creation.
        """
        self.nonstandardFrame = NonStandardShiftFrame(self,self.controller)
        self.standardFrame = StandardShiftFrame(self,self.controller)
        self.nonstandardFrame.grid(row=0, column=1, columnspan=2, rowspan=2, sticky="ne", pady=0, padx=4)
        self.standardFrame.grid(row=0,column=0,columnspan=1, rowspan=5, sticky="nw", pady=0)

        self.undo_button = tk.Button(self, text="Undo", command=self.controller.undo, width=13, height=2, foreground="#000000") 
        self.undo_button.grid(row=2, column=1, columnspan=1, sticky='e', pady=(10,0), padx=(0,1))

        self.redo_button = tk.Button(self, text="Redo", command=self.controller.redo, width=13, height=2, foreground="#000000")
        self.redo_button.grid(row=2, column=2, columnspan=1, sticky="w", pady=(10,0), padx=(1,0))

        self.swap_button = tk.Button(self, text="Swap", command=lambda: self.controller._perform_with_undo(lambda: self.controller.swap()), width=13, height=2, foreground="#000000")
        self.swap_button.grid(row=3, column=1, columnspan=2, sticky="w", pady=(10,0), padx=(5,0))

        self.balance_button = tk.Button(self, text="Balance", command=self.controller.show_balance_rules_dialog, width=13, height=2, foreground="#000000")
        self.balance_button.grid(row=3, column=2, columnspan=2, sticky="w", pady=(10,0), padx=(1,5))


class NonStandardShiftFrame(tk.Frame):
    def __init__(self, parent: inputFrame, controller: ScheduleApp):
        super().__init__(parent, relief=tk.GROOVE, borderwidth=2)
        self.parent = parent
        self.controller = controller
        self.create_widgets()

    def create_widgets(self):
        label = tk.Label(self, text=" Add nonstandard shift\n use DELETE to clear selection.",justify='left')
        label.grid(row=0, column=0, columnspan=3, sticky="w", pady=5)

        self.entry = tk.Entry(self, width=20)
        self.entry.grid(row=1, column=0, padx=5, pady=5)

        add_button = tk.Button(self, text="Add Shift", command=self.add_shift_action)
        add_button.grid(row=1, column=1, padx=5, pady=5)


    def add_shift_action(self):
        """Activates when 'Add Shift' button is pressed."""
        selection = self.controller.get_sheet_selection()
        if not selection:
            return
        shift = self.entry.get()
        if shift == 'DELETE': 
            shift = np.nan
        selection['shift'] = shift

        self.controller._perform_with_undo(self.controller.add_nonstandard_shift, selection)


class StandardShiftFrame(tk.Frame):
    def __init__(self, parent: inputFrame, controller: ScheduleApp):
        super().__init__(parent, relief=tk.GROOVE, borderwidth=2)
        self.parent = parent
        self.controller = controller
        self.create_widgets()

    def create_widgets(self):
        label = tk.Label(self, text="Add standard shift")
        label.grid(row=0, column=0, columnspan=3, sticky="w", pady=5)

        self.listbox = tk.Listbox(self, selectmode='single',width=20)
        self.listbox.grid(row=1, column=0, padx=5, pady=5)
        for shift in STANDARD_FLOOR_SHIFTS:
            self.listbox.insert(tk.END, shift)
            self.listbox.itemconfig(tk.END,bg=SHIFT_INFO[shift]['color'])


        add_button = tk.Button(self, text="Add Shift", command=self.add_standard_action, height=10)
        add_button.grid(row=1, column=1, padx=5, pady=5)

    def add_standard_action(self):
        """Activates when 'Add shift' button is pressed."""
        shift = self.listbox.get(tk.ACTIVE)
        self.controller._perform_with_undo(self.controller.add_standard_shift, shift)


def show_ui() -> None:
    """Creates a GUI."""
    app = ScheduleApp()
    app.protocol('WM_DELETE_WINDOW', app.close)
    app.mainloop()

if __name__ == '__main__':
    show_ui()
    '''

    Elements unfinished:
    - text document paging system. General, then each day of the week.
    - column width for sheetFrame. Add a dynamic width when sheet is created.

    Fun adds:
    - Shift balancer with adjustable priorities.
    - Separate file for shift and color information.
    - ability to set colors of shifts with CELL_COLOR values.
    - copy and paste functionality.
    '''
    #open_file(RES_FILE_NAME)